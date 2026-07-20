"""
routers/excel.py — Exportar e importar todos los datos en formato .xlsx
Usa openpyxl (sin dependencias externas de JS).
"""
import io
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.database import get_conn, CATEGORIAS

router = APIRouter(prefix="/excel", tags=["excel"])

# Colores del tema
COLOR_HEADER   = "1F3864"   # azul oscuro
COLOR_INGRESO  = "D9EAD3"   # verde claro
COLOR_GASTO    = "FCE4D6"   # rojo claro
COLOR_PRES     = "DEEAF1"   # azul claro
COLOR_META     = "FFF2CC"   # amarillo claro
COLOR_WHITE    = "FFFFFF"
FONT_HEADER    = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE     = Font(bold=True, size=13, color="1F3864")


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_row(ws, cols: list[str], fill_color: str):
    fill = PatternFill("solid", fgColor=fill_color)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col) + 4, 14)


def _data_row(ws, row_num: int, values: list, bg: Optional[str] = None):
    fill = PatternFill("solid", fgColor=bg) if bg else None
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.border = _border()
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill
        if isinstance(val, float):
            cell.number_format = '#,##0.00'


# ── EXPORT ────────────────────────────────────────────────────────────────────

@router.get("/exportar")
def exportar():
    """Descarga todos los datos como archivo .xlsx con 4 hojas."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # quita hoja default

    with get_conn() as conn:
        ingresos    = conn.execute("SELECT fecha,fuente,desc,monto FROM ingresos ORDER BY fecha DESC").fetchall()
        gastos      = conn.execute("SELECT fecha,cat,desc,monto,metodo FROM gastos ORDER BY fecha DESC").fetchall()
        presupuestos= conn.execute("SELECT cat,monto FROM presupuestos ORDER BY cat").fetchall()
        metas       = conn.execute("SELECT nombre,desc,objetivo,ahorrado,fecha FROM metas ORDER BY id").fetchall()

    # ── Hoja Ingresos
    ws = wb.create_sheet("Ingresos")
    _header_row(ws, ["Fecha", "Fuente / Persona", "Descripcion", "Monto ($)"], COLOR_HEADER)
    for i, r in enumerate(ingresos, 2):
        _data_row(ws, i, [r["fecha"], r["fuente"], r["desc"], r["monto"]],
                  COLOR_INGRESO if i % 2 == 0 else COLOR_WHITE)
    # Total
    ws.cell(row=len(ingresos)+2, column=3, value="TOTAL").font = Font(bold=True)
    tot = ws.cell(row=len(ingresos)+2, column=4, value=sum(r["monto"] for r in ingresos))
    tot.font = Font(bold=True)
    tot.number_format = '#,##0.00'

    # ── Hoja Gastos
    ws2 = wb.create_sheet("Gastos")
    _header_row(ws2, ["Fecha","Categoria","Descripcion","Monto ($)","Metodo de Pago"], COLOR_HEADER)
    for i, r in enumerate(gastos, 2):
        _data_row(ws2, i, [r["fecha"],r["cat"],r["desc"],r["monto"],r["metodo"]],
                  COLOR_GASTO if i % 2 == 0 else COLOR_WHITE)
    ws2.cell(row=len(gastos)+2, column=3, value="TOTAL").font = Font(bold=True)
    tot2 = ws2.cell(row=len(gastos)+2, column=4, value=sum(r["monto"] for r in gastos))
    tot2.font = Font(bold=True)
    tot2.number_format = '#,##0.00'

    # ── Hoja Presupuesto
    ws3 = wb.create_sheet("Presupuesto")
    _header_row(ws3, ["Categoria","Presupuesto Mensual ($)","Gastado ($)","Restante ($)","% Usado"], COLOR_HEADER)
    gastos_por_cat = {}
    for g in gastos:
        gastos_por_cat[g["cat"]] = gastos_por_cat.get(g["cat"], 0) + g["monto"]
    for i, r in enumerate(presupuestos, 2):
        pres   = r["monto"]
        gast   = gastos_por_cat.get(r["cat"], 0)
        rest   = pres - gast
        pct    = round(gast / pres * 100, 1) if pres > 0 else 0
        _data_row(ws3, i, [r["cat"], pres, gast, rest, pct],
                  COLOR_PRES if i % 2 == 0 else COLOR_WHITE)
        ws3.cell(row=i, column=5).number_format = '0.0"%"'

    # ── Hoja Metas
    ws4 = wb.create_sheet("Metas de Ahorro")
    _header_row(ws4, ["Meta","Descripcion","Objetivo ($)","Ahorrado ($)","Falta ($)","Fecha Objetivo","% Logrado"], COLOR_HEADER)
    for i, r in enumerate(metas, 2):
        falta = r["objetivo"] - r["ahorrado"]
        pct   = round(r["ahorrado"] / r["objetivo"] * 100, 1) if r["objetivo"] > 0 else 0
        _data_row(ws4, i, [r["nombre"],r["desc"],r["objetivo"],r["ahorrado"],falta,r["fecha"],pct],
                  COLOR_META if i % 2 == 0 else COLOR_WHITE)
        ws4.cell(row=i, column=7).number_format = '0.0"%"'

    # Stream como descarga
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="presupuesto_{fecha_hoy}.xlsx"'}
    )


# ── IMPORT ────────────────────────────────────────────────────────────────────

@router.post("/importar")
def importar(archivo: UploadFile = File(...)):
    """Lee un .xlsx exportado por esta app y reemplaza los datos."""
    if not archivo.filename.endswith(".xlsx"):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx")

    contenido = archivo.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Archivo Excel invalido: {e}")

    with get_conn() as conn:
        # Limpiar tablas
        conn.execute("DELETE FROM ingresos")
        conn.execute("DELETE FROM gastos")
        conn.execute("DELETE FROM metas")

        # Ingresos
        if "Ingresos" in wb.sheetnames:
            ws = wb["Ingresos"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha, fuente, desc, monto = (row + (None,)*4)[:4]
                if fuente and monto:
                    try:
                        conn.execute(
                            "INSERT INTO ingresos (fecha,fuente,desc,monto) VALUES (?,?,?,?)",
                            (str(fecha) if fecha else None, str(fuente), str(desc) if desc else None, float(monto))
                        )
                    except Exception:
                        pass

        # Gastos
        if "Gastos" in wb.sheetnames:
            ws2 = wb["Gastos"]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                fecha, cat, desc, monto, metodo = (row + (None,)*5)[:5]
                if cat and monto:
                    try:
                        conn.execute(
                            "INSERT INTO gastos (fecha,cat,desc,monto,metodo) VALUES (?,?,?,?,?)",
                            (str(fecha) if fecha else None, str(cat), str(desc) if desc else None,
                             float(monto), str(metodo) if metodo else "Efectivo")
                        )
                    except Exception:
                        pass

        # Presupuesto
        if "Presupuesto" in wb.sheetnames:
            ws3 = wb["Presupuesto"]
            for row in ws3.iter_rows(min_row=2, values_only=True):
                cat, monto = (row + (None, None))[:2]
                if cat and monto is not None:
                    try:
                        conn.execute(
                            "INSERT INTO presupuestos (cat,monto) VALUES (?,?) ON CONFLICT(cat) DO UPDATE SET monto=excluded.monto",
                            (str(cat), float(monto))
                        )
                    except Exception:
                        pass

        # Metas
        if "Metas de Ahorro" in wb.sheetnames:
            ws4 = wb["Metas de Ahorro"]
            for row in ws4.iter_rows(min_row=2, values_only=True):
                nombre, desc, objetivo, ahorrado, _, fecha, _ = (row + (None,)*7)[:7]
                if nombre and objetivo:
                    try:
                        conn.execute(
                            "INSERT INTO metas (nombre,desc,objetivo,ahorrado,fecha) VALUES (?,?,?,?,?)",
                            (str(nombre), str(desc) if desc else None,
                             float(objetivo), float(ahorrado or 0), str(fecha) if fecha else None)
                        )
                    except Exception:
                        pass

    return {"ok": True, "mensaje": "Datos importados correctamente desde Excel"}
